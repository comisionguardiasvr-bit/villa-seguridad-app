import streamlit as st
import pandas as pd
import json
import io
import time
from datetime import datetime
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from fpdf import FPDF
from streamlit_gsheets import GSheetsConnection
import os

# ==========================================
# ⚙️ CONFIGURACIÓN GLOBAL
# ==========================================
st.set_page_config(page_title="Tesorería Villa Raimapu", page_icon="🌿", layout="wide", initial_sidebar_state="expanded")

CUOTA_MENSUAL = 10000 
MESES_DISPONIBLES = ["Abril 2026", "Mayo 2026", "Junio 2026", "Julio 2026", "Agosto 2026", "Septiembre 2026", "Octubre 2026", "Noviembre 2026", "Diciembre 2026"]

# ==========================================
# 🎨 DISEÑO PROFESIONAL FINTECH (MODO CLARO/OSCURO)
# ==========================================
st.markdown("""
<style>
    /* Tipografía y acentos */
    h1, h2, h3, h4, h5 { font-family: 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; color: #5a7d65 !important; }
    
    /* Tarjetas de Métricas Ejecutivas */
    div[data-testid="metric-container"] {
        background-color: var(--secondary-background-color); 
        border: 1px solid rgba(90, 125, 101, 0.15);
        padding: 18px 20px; border-radius: 16px; 
        border-left: 6px solid #6c8d76; transition: all 0.3s ease;
        box-shadow: 0 4px 6px rgba(0,0,0,0.02);
    }
    div[data-testid="metric-container"]:hover { transform: translateY(-4px); box-shadow: 0 8px 15px rgba(0,0,0,0.08); }
    
    /* Botones Modernos */
    .stButton>button {
        border-radius: 10px !important; font-weight: 600 !important; letter-spacing: 0.5px;
        background-color: #6c8d76 !important; color: white !important; 
        border: none !important; transition: all 0.3s cubic-bezier(.25,.8,.25,1);
    }
    .stButton>button:hover { background-color: #4a6652 !important; transform: scale(1.02); box-shadow: 0 4px 10px rgba(74, 102, 82, 0.3); }
    
    /* Pestañas limpias */
    .stTabs [data-baseweb="tab-list"] { gap: 10px; border-bottom: 2px solid rgba(90, 125, 101, 0.1); }
    .stTabs [data-baseweb="tab"] { border-radius: 8px 8px 0px 0px; padding: 12px 16px; transition: background-color 0.3s; }
    .stTabs [data-baseweb="tab"]:hover { background-color: rgba(90, 125, 101, 0.05); }
    .stTabs [aria-selected="true"] { border-bottom: 3px solid #5a7d65 !important; color: #5a7d65 !important; font-weight: 600;}
    
    /* Formularios y Cajas */
    div[data-testid="stForm"], .stAlert { 
        border-radius: 16px !important; border: 1px solid rgba(90, 125, 101, 0.2) !important; 
    }
    
    #MainMenu, footer {visibility: hidden;}
    [data-testid="stImage"] img { border-radius: 50%; box-shadow: 0 4px 15px rgba(0,0,0,0.1); border: 4px solid #fdfcf9;}
</style>
""", unsafe_allow_html=True)

def fmt_dinero(monto):
    return f"$ {int(monto):,.0f}".replace(",", ".")

# ==========================================
# 🔐 SISTEMA DE AUTENTICACIÓN MULTI-ROL
# ==========================================
if 'autenticado' not in st.session_state:
    st.session_state.update({'autenticado': False, 'rol': "", 'usuario': ""})

if not st.session_state.autenticado:
    st.markdown("<br>", unsafe_allow_html=True)
    _, col_logo, _ = st.columns([2.5, 1.5, 2.5])
    with col_logo:
        if os.path.exists("logo_villa.jpg"): st.image("logo_villa.jpg", use_container_width=True)
        else: st.markdown("<h1 style='text-align: center; font-size: 4em;'>🌿</h1>", unsafe_allow_html=True)
    
    st.markdown("<h2 style='text-align: center;'>Plataforma Financiera Raimapu</h2>", unsafe_allow_html=True)
    st.markdown("<p style='text-align: center; color: gray;'>Gestión transparente y segura</p><br>", unsafe_allow_html=True)
    
    _, col_login, _ = st.columns([1, 1.5, 1])
    with col_login:
        with st.form("login_form"):
            tipo_usuario = st.selectbox("Seleccione su perfil de acceso:", ["Tesorera / Administradora", "Recaudadora 1", "Recaudadora 2", "Recaudadora 3"])
            pass_input = st.text_input("🔑 Contraseña de seguridad:", type="password")
            if st.form_submit_button("Ingresar de forma segura", use_container_width=True):
                # Validaciones
                credenciales = {
                    "Tesorera / Administradora": ("villa2026", "Admin", "Tesorera Principal"),
                    "Recaudadora 1": ("recauda1", "Recaudadora", "Recaudadora 1"),
                    "Recaudadora 2": ("recauda2", "Recaudadora", "Recaudadora 2"),
                    "Recaudadora 3": ("recauda3", "Recaudadora", "Recaudadora 3")
                }
                
                if tipo_usuario in credenciales and pass_input == credenciales[tipo_usuario][0]:
                    st.session_state.autenticado = True
                    st.session_state.rol = credenciales[tipo_usuario][1]
                    st.session_state.usuario = credenciales[tipo_usuario][2]
                    st.rerun()
                else:
                    st.error("❌ Credenciales incorrectas. Acceso denegado.")
    st.stop()

# ==========================================
# 🛠️ CONEXIÓN Y TRANSACCIONES SEGURAS (ANTI-BLOQUEOS)
# ==========================================
conn = st.connection("gsheets", type=GSheetsConnection)

def cargar_hoja_robusta(nombre_hoja, columnas_esperadas):
    try:
        df = conn.read(worksheet=nombre_hoja, ttl=0).dropna(how="all")
        if df.empty: return pd.DataFrame(columns=columnas_esperadas)
        for col in columnas_esperadas:
            if col not in df.columns: df[col] = None
        return df
    except Exception: 
        return pd.DataFrame(columns=columnas_esperadas)

def ejecutar_transaccion(worksheet_name, dataframe_actualizado, cache_func, msg_exito="Operación exitosa"):
    """Manejador de transacciones seguro para evitar cuelgues de API."""
    try:
        conn.update(worksheet=worksheet_name, data=dataframe_actualizado)
        cache_func.clear()
        st.toast(f"✅ {msg_exito}")
        return True
    except Exception as e:
        st.error("⚠️ Error de sincronización con Google (Servidor ocupado). Por favor, espere 5 segundos y vuelva a intentarlo.")
        return False

@st.cache_data(ttl=600, show_spinner=False)
def cargar_pagos():
    df = cargar_hoja_robusta("Pagos", ['calle', 'numero', 'propietario', 'monto_pagado', 'fecha', 'mes', 'registrado_por', 'metodo_pago'])
    if not df.empty: df['numero'] = pd.to_numeric(df['numero'], errors='coerce').fillna(0).astype(int)
    return df

@st.cache_data(ttl=600, show_spinner=False)
def cargar_gastos(): return cargar_hoja_robusta("Gastos", ['descripcion', 'monto', 'fecha', 'mes'])

@st.cache_data(ttl=600, show_spinner=False)
def cargar_extra(): return cargar_hoja_robusta("Ingresos_Extra", ['concepto', 'monto', 'fecha', 'mes'])

@st.cache_data(ttl=600, show_spinner=False)
def cargar_guardias(): return cargar_hoja_robusta("Guardias", ['nombre', 'tipo', 'sueldo'])

@st.cache_data(ttl=600, show_spinner=False)
def cargar_logs(): return cargar_hoja_robusta("Logs", ['fecha_hora', 'usuario', 'accion', 'detalle'])

@st.cache_data(ttl=600, show_spinner=False)
def cargar_ajustes(): return cargar_hoja_robusta("Ajustes_Guardias", ['mes', 'guardia', 'tipo', 'monto', 'detalle'])

@st.cache_data
def cargar_casas():
    with open('casas.json', 'r', encoding='utf-8') as f:
        df = pd.DataFrame(json.load(f))
        df['numero'] = pd.to_numeric(df['numero']).astype(int)
        return df.sort_values(by=['calle', 'numero']).reset_index(drop=True)

df_pagos_full = cargar_pagos()
df_gastos_full = cargar_gastos()
df_extra_full = cargar_extra()
df_guardias = cargar_guardias()
df_logs_full = cargar_logs()
df_ajustes_full = cargar_ajustes()
df_casas = cargar_casas()

def registrar_log(accion, detalle):
    df_l = cargar_logs()
    nuevo_log = pd.DataFrame([{'fecha_hora': datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 'usuario': st.session_state.usuario, 'accion': accion, 'detalle': detalle}])
    try:
        conn.update(worksheet="Logs", data=pd.concat([df_l, nuevo_log], ignore_index=True))
        cargar_logs.clear()
    except: pass # Los logs fallidos no deben detener la aplicación principal

# --- MENÚ LATERAL ---
with st.sidebar:
    if os.path.exists("logo_villa.jpg"): st.image("logo_villa.jpg", width=150)
    st.markdown(f"<h3 style='text-align: center;'>👤 {st.session_state.usuario}</h3>", unsafe_allow_html=True)
    st.markdown("---")
    mes_actual = st.selectbox("📅 Mes Operativo:", MESES_DISPONIBLES)
    st.markdown("---")
    if st.button("🚪 Cerrar Sesión", use_container_width=True):
        st.session_state.clear(); st.rerun()

def filtrar_por_mes(df, mes):
    if not df.empty and 'mes' in df.columns: return df[df['mes'].astype(str).str.lower() == mes.lower()].reset_index(drop=True)
    return pd.DataFrame(columns=df.columns)

df_pagos_mes = filtrar_por_mes(df_pagos_full, mes_actual)
df_gastos_mes = filtrar_por_mes(df_gastos_full, mes_actual)
df_extra_mes = filtrar_por_mes(df_extra_full, mes_actual)
df_ajustes_mes = filtrar_por_mes(df_ajustes_full, mes_actual)

# ==========================================
# 💰 MATEMÁTICA Y ARRASTRE VECTORIZADO
# ==========================================
idx_mes = MESES_DISPONIBLES.index(mes_actual)
meses_anteriores = MESES_DISPONIBLES[:idx_mes]

ingresos_ant = pd.to_numeric(df_pagos_full[df_pagos_full['mes'].isin(meses_anteriores)]['monto_pagado'], errors='coerce').fillna(0).sum() + \
               pd.to_numeric(df_extra_full[df_extra_full['mes'].isin(meses_anteriores)]['monto'], errors='coerce').fillna(0).sum() if idx_mes > 0 else 0

egresos_ant = pd.to_numeric(df_gastos_full[df_gastos_full['mes'].isin(meses_anteriores)]['monto'], errors='coerce').fillna(0).sum() if idx_mes > 0 else 0

if idx_mes > 0:
    for m_ant in meses_anteriores:
        sueldos_base = pd.to_numeric(df_guardias['sueldo'], errors='coerce').fillna(0).sum() if not df_guardias.empty else 0
        bonos = pd.to_numeric(df_ajustes_full[(df_ajustes_full['mes'] == m_ant) & (df_ajustes_full['tipo'] == 'Bono Turno Extra')]['monto'], errors='coerce').fillna(0).sum()
        descuentos = pd.to_numeric(df_ajustes_full[(df_ajustes_full['mes'] == m_ant) & (df_ajustes_full['tipo'] == 'Descuento Falta')]['monto'], errors='coerce').fillna(0).sum()
        egresos_ant += (sueldos_base + bonos - descuentos)

caja_chica_anterior = ingresos_ant - egresos_ant

ingresos_vecinos = pd.to_numeric(df_pagos_mes['monto_pagado'], errors='coerce').fillna(0).sum()
ingresos_eventos = pd.to_numeric(df_extra_mes['monto'], errors='coerce').fillna(0).sum()
total_ingresos_mes = ingresos_vecinos + ingresos_eventos

sueldos_base_actual = pd.to_numeric(df_guardias['sueldo'], errors='coerce').fillna(0).sum()
bonos_actual = pd.to_numeric(df_ajustes_mes[df_ajustes_mes['tipo'] == 'Bono Turno Extra']['monto'], errors='coerce').fillna(0).sum()
descuentos_actual = pd.to_numeric(df_ajustes_mes[df_ajustes_mes['tipo'] == 'Descuento Falta']['monto'], errors='coerce').fillna(0).sum()
egresos_guardias = sueldos_base_actual + bonos_actual - descuentos_actual

egresos_otros = pd.to_numeric(df_gastos_mes['monto'], errors='coerce').fillna(0).sum()
total_egresos_mes = egresos_guardias + egresos_otros

balance_final = caja_chica_anterior + total_ingresos_mes - total_egresos_mes
deudores_count = len(df_casas) - len(df_pagos_mes)

# ==========================================
# 📄 GENERADORES DE DOCUMENTOS
# ==========================================
def generar_pdf_cierre(ing_v, ing_e, egr_g, egr_o, bal, arrastre, df_ex, mes_t):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", 'B', 14)
    pdf.cell(0, 8, txt="BALANCE MENSUAL DE TESORERIA", ln=True, align='L')
    pdf.set_font("Arial", 'B', 10)
    pdf.cell(0, 6, txt="VILLA RAIMAPU - TIERRA FLORIDA - PUENTE ALTO", ln=True, align='L')
    pdf.set_font("Arial", '', 10)
    pdf.cell(0, 6, txt=f"Mes: {mes_t.upper()} | Generado el: {datetime.now().strftime('%d/%m/%Y')}", ln=True)
    pdf.ln(8)
    
    pdf.set_fill_color(235, 240, 235)
    pdf.set_font("Arial", 'B', 12)
    pdf.cell(0, 8, txt="1. Resumen de Caja", ln=True, fill=True)
    pdf.set_font("Arial", '', 11)
    if arrastre != 0:
        pdf.cell(120, 8, txt="(+) Saldo a favor mes anterior (Caja Chica)", border=1); pdf.cell(70, 8, txt=fmt_dinero(arrastre), border=1, ln=True, align='R')
    pdf.cell(120, 8, txt="(+) Recaudacion Cuotas Vecinos", border=1); pdf.cell(70, 8, txt=fmt_dinero(ing_v), border=1, ln=True, align='R')
    pdf.cell(120, 8, txt="(+) Ingresos Extra (Eventos/Rifas)", border=1); pdf.cell(70, 8, txt=fmt_dinero(ing_e), border=1, ln=True, align='R')
    pdf.cell(120, 8, txt="(-) Liquidaciones Guardias (Base + Ajustes)", border=1); pdf.cell(70, 8, txt=f"- {fmt_dinero(egr_g)}", border=1, ln=True, align='R')
    pdf.cell(120, 8, txt="(-) Gastos Operativos e Insumos", border=1); pdf.cell(70, 8, txt=f"- {fmt_dinero(egr_o)}", border=1, ln=True, align='R')
    pdf.set_font("Arial", 'B', 11)
    pdf.cell(120, 10, txt="SALDO FINAL DISPONIBLE EN CAJA", border=1); pdf.cell(70, 10, txt=fmt_dinero(bal), border=1, ln=True, align='R')
    
    if not df_ex.empty:
        pdf.ln(10); pdf.set_font("Arial", 'B', 12); pdf.cell(0, 8, txt="2. Detalle de Ingresos Extra", ln=True, fill=True)
        pdf.set_font("Arial", 'B', 10); pdf.cell(120, 8, txt="Actividad / Concepto", border=1); pdf.cell(70, 8, txt="Monto Recaudado", border=1, ln=True, align='C')
        pdf.set_font("Arial", '', 10)
        for _, r in df_ex.iterrows():
            m_val = pd.to_numeric(r['monto'], errors='coerce')
            pdf.cell(120, 8, txt=str(r['concepto']), border=1); pdf.cell(70, 8, txt=fmt_dinero(m_val) if not pd.isna(m_val) else "$ 0", border=1, ln=True, align='R')
            
    return pdf.output(dest='S').encode('latin-1', 'replace')

def generar_liquidacion_guardia(nombre, tipo, base, bonos, descuentos, total, mes_texto):
    pdf = FPDF(format='A5')
    pdf.add_page()
    pdf.set_font("Arial", 'B', 12)
    pdf.cell(0, 8, txt="LIQUIDACION DE PAGO - SERVICIO DE SEGURIDAD", ln=True, align='C')
    pdf.set_font("Arial", 'B', 9)
    pdf.cell(0, 5, txt="VILLA RAIMAPU - TIERRA FLORIDA", ln=True, align='C')
    pdf.ln(8)
    
    pdf.set_font("Arial", 'B', 10)
    pdf.cell(40, 6, txt="Nombre Trabajador:", border=0); pdf.set_font("Arial", '', 10); pdf.cell(0, 6, txt=nombre, ln=True)
    pdf.set_font("Arial", 'B', 10)
    pdf.cell(40, 6, txt="Periodo / Mes:", border=0); pdf.set_font("Arial", '', 10); pdf.cell(0, 6, txt=mes_texto.upper(), ln=True)
    pdf.set_font("Arial", 'B', 10)
    pdf.cell(40, 6, txt="Tipo de Turno:", border=0); pdf.set_font("Arial", '', 10); pdf.cell(0, 6, txt=tipo, ln=True)
    pdf.ln(5)
    
    pdf.set_font("Arial", 'B', 10)
    pdf.cell(90, 8, txt="Detalle de Remuneracion", border=1); pdf.cell(40, 8, txt="Monto", border=1, ln=True, align='C')
    pdf.set_font("Arial", '', 10)
    pdf.cell(90, 8, txt="Sueldo Base Acordado", border=1); pdf.cell(40, 8, txt=fmt_dinero(base), border=1, ln=True, align='R')
    if bonos > 0:
        pdf.cell(90, 8, txt="(+) Bonos por Turnos Extra", border=1); pdf.cell(40, 8, txt=fmt_dinero(bonos), border=1, ln=True, align='R')
    if descuentos > 0:
        pdf.cell(90, 8, txt="(-) Descuentos por Inasistencias", border=1); pdf.cell(40, 8, txt=f"- {fmt_dinero(descuentos)}", border=1, ln=True, align='R')
    
    pdf.set_font("Arial", 'B', 12)
    pdf.cell(90, 10, txt="TOTAL A PAGAR LIQUIDO", border=1); pdf.cell(40, 10, txt=fmt_dinero(total), border=1, ln=True, align='R')
    
    pdf.ln(15)
    pdf.set_font("Arial", '', 9)
    pdf.cell(0, 5, txt="Recibi conforme el pago total de mis servicios correspondientes al mes indicado.", ln=True, align='C')
    pdf.ln(15)
    pdf.cell(0, 5, txt="_________________________________________", ln=True, align='C')
    pdf.cell(0, 5, txt=f"Firma: {nombre}", ln=True, align='C')
    
    return pdf.output(dest='S').encode('latin-1', 'replace')

def generar_excel_morosos(df_morosos, mes_texto):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as wr:
        df_export = df_morosos[['calle', 'numero', 'propietario']].copy()
        df_export.columns = [f"VECINOS MOROSOS - {mes_texto.upper()}", "N° CASA", "PROPIETARIO"]
        
        df_export.to_excel(wr, index=False, sheet_name='Morosos')
        ws = wr.sheets['Morosos']
        
        azul_claro = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
        fuente_titulos = Font(name='Calibri', size=16, bold=True)
        fuente_datos = Font(name='Calibri', size=11)
        centrado = Alignment(horizontal="center", vertical="center")
        borde_fino = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        ws.column_dimensions['A'].width = 38
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 35
        
        for col in range(1, 4):
            cell_header = ws.cell(row=1, column=col)
            cell_header.fill = azul_claro
            cell_header.font = fuente_titulos
            cell_header.alignment = centrado
            cell_header.border = borde_fino
            
            for row in range(2, len(df_export) + 2):
                cell_data = ws.cell(row=row, column=col)
                cell_data.font = fuente_datos
                cell_data.alignment = centrado
                cell_data.border = borde_fino
                
    return buf.getvalue()

# ==========================================
# INTERFAZ: RECAUDADORA EN TERRENO
# ==========================================
if st.session_state.rol == "Recaudadora":
    st.markdown(f"<h2>📱 Portal de Recaudación</h2>", unsafe_allow_html=True)
    st.info(f"💡 Bienvenida **{st.session_state.usuario}**. Usa este panel para registrar cobros en terreno.")
    
    with st.container():
        c_sel = st.selectbox("1. Seleccione Pasaje", df_casas['calle'].unique())
        p_ya = df_pagos_mes[df_pagos_mes['calle'] == c_sel]['numero'].tolist()
        pend = df_casas[(df_casas['calle'] == c_sel) & (~df_casas['numero'].isin(p_ya))]
        
        if pend.empty: 
            st.success("🎉 ¡Todas las casas de este pasaje están al día este mes!")
        else:
            opc = {f"N° {r['numero']} - {r['propietario']}": r['numero'] for _, r in pend.iterrows()}
            n_sel = opc[st.selectbox("2. Seleccione Casa", opc.keys())]
            nom_vecino = df_casas[(df_casas['calle'] == c_sel) & (df_casas['numero'] == n_sel)]['propietario'].values[0]
            
            meses_a_pagar = st.multiselect("3. Meses que está pagando:", MESES_DISPONIBLES, default=[mes_actual])
            metodo_pago = st.selectbox("4. Método de Pago:", ["Efectivo", "Transferencia Bancaria", "Otro"])
            
            total_cobrar = len(meses_a_pagar) * CUOTA_MENSUAL
            st.markdown(f"<h3 style='text-align: center; color: #d35400;'>Total a Recibir: {fmt_dinero(total_cobrar)}</h3>", unsafe_allow_html=True)
            
            if st.button("💳 Confirmar y Registrar Pago", type="primary", use_container_width=True):
                nuevos_pagos = []
                fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M")
                for m in meses_a_pagar:
                    nuevos_pagos.append({
                        'calle': c_sel, 'numero': int(n_sel), 'propietario': nom_vecino, 
                        'monto_pagado': CUOTA_MENSUAL, 'fecha': fecha_actual, 'mes': m, 
                        'registrado_por': st.session_state.usuario, 'metodo_pago': metodo_pago
                    })
                
                # Transacción Segura
                df_act = pd.concat([df_pagos_full, pd.DataFrame(nuevos_pagos)], ignore_index=True)
                exito = ejecutar_transaccion("Pagos", df_act, cargar_pagos, "Pagos guardados en la base central.")
                
                if exito:
                    registrar_log("Cobro", f"Cobró {fmt_dinero(total_cobrar)} a {c_sel} #{n_sel} en {metodo_pago}")
                    
                    # Generador de Comprobante para WhatsApp
                    meses_str = ", ".join(meses_a_pagar)
                    texto_wp = f"✅ *COMPROBANTE DE PAGO VILLA RAIMAPU*\n\nVecino/a: {nom_vecino}\nCasa: {c_sel} #{n_sel}\nMeses Pagados: {meses_str}\nMonto Total: {fmt_dinero(total_cobrar)}\nMétodo: {metodo_pago}\nFecha: {fecha_actual}\n\n_Recibido por: {st.session_state.usuario}_\n🌿 ¡Gracias por su compromiso con la seguridad de nuestra villa!"
                    
                    st.success("Operación completada. Copie el mensaje a continuación para enviarlo por WhatsApp al vecino:")
                    st.code(texto_wp, language="text")
                    
    st.markdown("---")
    st.warning("⚠️ **¿Te equivocaste anotando un pago?** No te preocupes. Avísale a la Tesorera indicando la casa y ella lo anulará en el sistema central para mantener la cuadratura intacta.")

# ==========================================
# INTERFAZ: ADMINISTRADORA / TESORERÍA
# ==========================================
elif st.session_state.rol == "Admin":
    st.markdown(f"<h2>🏢 Panel de Tesorería General</h2>", unsafe_allow_html=True)
    
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Caja Chica Mes Anterior", fmt_dinero(caja_chica_anterior))
    m2.metric(f"Nuevos Ingresos ({mes_actual})", fmt_dinero(total_ingresos_mes))
    m3.metric("Fondo Total Disponible", fmt_dinero(balance_final))
    m4.metric("Casas Morosas del Mes", f"{deudores_count} pendientes")
    
    t1, t2, t3, t4, t5 = st.tabs(["📝 1. Visor de Pagos", "🎁 2. Ingresos Extra", "🛒 3. Gastos", "👮‍♂️ 4. Personal", "📑 5. Documentos y Auditoría"])

    # ------------------ PESTAÑA 1: PAGOS ------------------
    with t1:
        st.markdown("#### 📋 Visor Global de Recaudación")
        st.write("Monitoree los cobros realizados por el equipo de terreno.")
        
        filtro_mes = st.selectbox("Filtrar la tabla por mes:", ["Ver todos los meses"] + MESES_DISPONIBLES)
        
        if not df_pagos_full.empty:
            df_mostrar = df_pagos_full.copy()
            if filtro_mes != "Ver todos los meses":
                df_mostrar = df_mostrar[df_mostrar['mes'] == filtro_mes]
            
            if not df_mostrar.empty:
                df_mostrar['monto_pagado'] = df_mostrar['monto_pagado'].apply(fmt_dinero)
                st.dataframe(df_mostrar[['calle', 'numero', 'propietario', 'mes', 'monto_pagado', 'metodo_pago', 'fecha', 'registrado_por']], use_container_width=True, hide_index=True)
            else:
                st.info(f"No hay pagos registrados para el mes de {filtro_mes}.")
        else:
            st.info("Aún no hay pagos registrados en la base de datos.")

        st.markdown("---")
        st.markdown("#### 🛠️ Herramienta de Corrección (Anular Pago)")
        st.write("Seleccione el registro exacto reportado por la recaudadora para eliminarlo de la base de datos.")
        
        if not df_pagos_full.empty:
            with st.expander("Abrir panel de anulación de registros"):
                with st.form("form_anular"):
                    st.warning("⚠️ Esta acción restará el dinero de la caja y generará una alerta en Auditoría.")
                    opciones_borrar = {f"ID {idx} | {r['calle']} #{r['numero']} ({r['mes']}) | {fmt_dinero(r['monto_pagado'])} | {r['registrado_por']}": idx for idx, r in df_pagos_full.iterrows()}
                    sel_str = st.selectbox("Seleccione el pago a eliminar:", list(opciones_borrar.keys()))
                    
                    if st.form_submit_button("Eliminar Pago Seleccionado", use_container_width=True):
                        idx_to_delete = opciones_borrar[sel_str]
                        dato = df_pagos_full.loc[idx_to_delete]
                        df_act = df_pagos_full.drop(idx_to_delete).reset_index(drop=True)
                        if ejecutar_transaccion("Pagos", df_act, cargar_pagos, "Pago anulado correctamente."):
                            registrar_log("Anulacion Pago", f"Eliminó pago ID:{idx_to_delete} de {dato['calle']} #{dato['numero']} ({dato['mes']})")
                            st.rerun()

    # ------------------ PESTAÑA 2: INGRESOS EXTRA ------------------
    with t2:
        st.markdown("#### 🎁 Ingresos Extraordinarios")
        with st.form("form_extra"):
            con, mon = st.text_input("Concepto (Rifas, Aportes Voluntarios)"), st.number_input("Monto Recaudado ($)", step=1000)
            if st.form_submit_button("💰 Ingresar a Caja", use_container_width=True) and con and mon > 0:
                nuevo_e = pd.DataFrame([{'concepto': con, 'monto': int(mon), 'fecha': datetime.now().strftime("%d/%m/%Y"), 'mes': mes_actual}])
                df_act = pd.concat([df_extra_full, nuevo_e], ignore_index=True)
                if ejecutar_transaccion("Ingresos_Extra", df_act, cargar_extra, "Ingreso registrado."):
                    registrar_log("Ingreso Extra", f"Agregó {fmt_dinero(mon)} por {con}")
                    st.rerun()
                    
        if not df_extra_mes.empty: 
            df_e_disp = df_extra_mes.copy()
            df_e_disp['monto'] = df_e_disp['monto'].apply(fmt_dinero)
            st.dataframe(df_e_disp[['concepto', 'monto', 'fecha']], use_container_width=True, hide_index=True)
            with st.expander("❌ Anular Ingreso Extra"):
                borrar_e = st.selectbox("Seleccione registro a borrar", [f"ID {idx} | {r['concepto']} - {fmt_dinero(r['monto'])}" for idx, r in df_extra_mes.iterrows()])
                if st.button("Confirmar Eliminación de Ingreso"):
                    idx_del = int(borrar_e.split(" | ")[0].replace("ID ", ""))
                    df_act = df_extra_full.drop(df_extra_full[df_extra_full.index == idx_del].index).reset_index(drop=True)
                    if ejecutar_transaccion("Ingresos_Extra", df_act, cargar_extra, "Registro eliminado."):
                        registrar_log("Borro Ingreso", f"Eliminó ID:{idx_del}")
                        st.rerun()

    # ------------------ PESTAÑA 3: GASTOS ------------------
    with t3:
        st.markdown("#### 🛒 Gastos Operativos de la Villa")
        with st.form("form_gastos"):
            des, val = st.text_input("Descripción del Gasto o Insumo"), st.number_input("Costo Total ($)", step=1000)
            if st.form_submit_button("🚀 Registrar Gasto", use_container_width=True) and des and val > 0:
                nuevo_g = pd.DataFrame([{'descripcion': des, 'monto': int(val), 'fecha': datetime.now().strftime("%d/%m/%Y"), 'mes': mes_actual}])
                df_act = pd.concat([df_gastos_full, nuevo_g], ignore_index=True)
                if ejecutar_transaccion("Gastos", df_act, cargar_gastos, "Gasto descontado de caja."):
                    registrar_log("Nuevo Gasto", f"Gastó {fmt_dinero(val)} en {des}")
                    st.rerun()
                    
        if not df_gastos_mes.empty: 
            df_g_disp = df_gastos_mes.copy()
            df_g_disp['monto'] = df_g_disp['monto'].apply(fmt_dinero)
            st.dataframe(df_g_disp[['descripcion', 'monto', 'fecha']], use_container_width=True, hide_index=True)
            with st.expander("❌ Anular Gasto"):
                borrar_g = st.selectbox("Seleccione gasto a borrar", [f"ID {idx} | {r['descripcion']} - {fmt_dinero(r['monto'])}" for idx, r in df_gastos_mes.iterrows()])
                if st.button("Confirmar Eliminación de Gasto"):
                    idx_del = int(borrar_g.split(" | ")[0].replace("ID ", ""))
                    df_act = df_gastos_full.drop(df_gastos_full[df_gastos_full.index == idx_del].index).reset_index(drop=True)
                    if ejecutar_transaccion("Gastos", df_act, cargar_gastos, "Registro eliminado."):
                        registrar_log("Borro Gasto", f"Eliminó ID:{idx_del}")
                        st.rerun()

    # ------------------ PESTAÑA 4: PERSONAL ------------------
    with t4:
        st.markdown("#### 👮‍♂️ Gestión de Recursos Humanos")
        c_ver, c_mod = st.columns(2)
        with c_ver:
            st.markdown("##### 💵 Nómina de Contratos")
            if not df_guardias.empty:
                df_g_disp = df_guardias.copy()
                df_g_disp['sueldo'] = df_g_disp['sueldo'].apply(fmt_dinero)
                st.dataframe(df_g_disp, use_container_width=True, hide_index=True)
            else:
                st.info("Sin personal registrado.")
                
            with st.expander("➕ Contratar Nuevo Personal"):
                with st.form("add_gu"):
                    n_gu = st.text_input("Nombre Completo")
                    t_gu = st.selectbox("Tipo de Turno", ["Full Time", "Part Time", "Reemplazo"])
                    s_gu = st.number_input("Sueldo Base ($)", value=400000, step=10000)
                    if st.form_submit_button("Guardar Contrato", use_container_width=True) and n_gu:
                        n_reg = pd.DataFrame([{'nombre': n_gu, 'tipo': t_gu, 'sueldo': int(s_gu)}])
                        df_act = pd.concat([df_guardias, n_reg], ignore_index=True)
                        if ejecutar_transaccion("Guardias", df_act, cargar_guardias, "Guardia contratado."):
                            registrar_log("Contrato Guardia", f"Ingresó a {n_gu} ({fmt_dinero(s_gu)})")
                            st.rerun()
                            
            if not df_guardias.empty:
                with st.expander("➖ Desvincular Personal"):
                    with st.form("del_gu"):
                        g_borrar = st.selectbox("Seleccione empleado", df_guardias['nombre'].tolist())
                        if st.form_submit_button("Confirmar Desvinculación", use_container_width=True):
                            df_act = df_guardias[df_guardias['nombre'] != g_borrar].reset_index(drop=True)
                            if ejecutar_transaccion("Guardias", df_act, cargar_guardias, "Guardia eliminado."):
                                registrar_log("Desvinculo Guardia", f"Eliminó a {g_borrar}")
                                st.rerun()

        with c_mod:
            st.markdown("##### ⚖️ Ajustar Liquidación (Mes Actual)")
            if not df_guardias.empty:
                with st.form("form_ajustes"):
                    g_sel = st.selectbox("Trabajador:", df_guardias['nombre'].tolist())
                    tipo_aj = st.radio("Tipo de Novedad:", ["Bono Turno Extra", "Descuento Falta"])
                    monto_aj = st.number_input("Monto a aplicar ($)", min_value=0, step=5000)
                    det_aj = st.text_input("Motivo o Fecha")
                    if st.form_submit_button("Guardar Novedad", use_container_width=True) and monto_aj > 0:
                        n_ajuste = pd.DataFrame([{'mes': mes_actual, 'guardia': g_sel, 'tipo': tipo_aj, 'monto': int(monto_aj), 'detalle': det_aj}])
                        df_act = pd.concat([df_ajustes_full, n_ajuste], ignore_index=True)
                        if ejecutar_transaccion("Ajustes_Guardias", df_act, cargar_ajustes, "Ajuste aplicado al sueldo."):
                            registrar_log("Ajuste Sueldo", f"Aplicó {tipo_aj} de {fmt_dinero(monto_aj)} a {g_sel}")
                            st.rerun()
                            
        if not df_ajustes_mes.empty:
            st.markdown("##### 📝 Historial de Novedades del mes:")
            df_aj_disp = df_ajustes_mes.copy()
            df_aj_disp['monto'] = df_aj_disp['monto'].apply(fmt_dinero)
            st.dataframe(df_aj_disp[['guardia', 'tipo', 'monto', 'detalle']], use_container_width=True, hide_index=True)

    # ------------------ PESTAÑA 5: REPORTES ------------------
    with t5:
        st.markdown("#### 📑 Central de Documentos y Auditoría")
        
        col_doc1, col_doc2 = st.columns(2)
        with col_doc1:
            st.markdown("##### 📊 Balance de Tesorería")
            doc_balance = generar_pdf_cierre(ingresos_vecinos, ingresos_eventos, egresos_guardias, egresos_otros, balance_final, caja_chica_anterior, df_extra_mes, mes_actual)
            st.download_button("📄 Descargar PDF Oficial", doc_balance, file_name=f"Balance_Raimapu_{mes_actual}.pdf", type="primary", use_container_width=True)
            
        with col_doc2:
            st.markdown("##### 📝 Control de Cobranza")
            df_deu = df_casas.merge(df_pagos_mes[['calle', 'numero']], on=['calle', 'numero'], how='left', indicator=True)
            df_deu = df_deu[df_deu['_merge'] == 'left_only'].drop(columns=['_merge'])
            excel_morosos = generar_excel_morosos(df_deu, mes_actual)
            st.download_button("📥 Descargar Morosos (Excel)", excel_morosos, file_name=f"Morosos_{mes_actual}.xlsx", type="primary", use_container_width=True)

        st.markdown("---")
        st.markdown("##### 👮‍♂️ Liquidaciones de Sueldo Trabajadores")
        if not df_guardias.empty:
            c_liq1, c_liq2 = st.columns([2, 1])
            with c_liq1:
                guardia_liq = st.selectbox("Seleccionar trabajador para emitir su recibo:", df_guardias['nombre'].tolist())
            with c_liq2:
                st.markdown("<br>", unsafe_allow_html=True)
                datos_g = df_guardias[df_guardias['nombre'] == guardia_liq].iloc[0]
                base = int(pd.to_numeric(datos_g['sueldo'], errors='coerce'))
                bonos = int(pd.to_numeric(df_ajustes_mes[(df_ajustes_mes['guardia'] == guardia_liq) & (df_ajustes_mes['tipo'] == 'Bono Turno Extra')]['monto'], errors='coerce').sum())
                descuentos = int(pd.to_numeric(df_ajustes_mes[(df_ajustes_mes['guardia'] == guardia_liq) & (df_ajustes_mes['tipo'] == 'Descuento Falta')]['monto'], errors='coerce').sum())
                total_pagar = base + bonos - descuentos
                
                pdf_liq = generar_liquidacion_guardia(guardia_liq, datos_g['tipo'], base, bonos, descuentos, total_pagar, mes_actual)
                st.download_button("📄 Imprimir Liquidación", pdf_liq, file_name=f"Liquidacion_{guardia_liq.replace(' ', '_')}_{mes_actual}.pdf", use_container_width=True)

        st.markdown("---")
        st.markdown("##### 🕵️‍♂️ Panel de Auditoría (Logs de Seguridad)")
        st.dataframe(df_logs_full.tail(20).sort_index(ascending=False), use_container_width=True, hide_index=True)
        
        buf_logs = io.BytesIO()
        with pd.ExcelWriter(buf_logs, engine='openpyxl') as wr:
            df_logs_full.to_excel(wr, index=False, sheet_name='Auditoria')
        st.download_button("📥 Descargar Historial Completo (Excel)", buf_logs.getvalue(), file_name=f"Auditoria_Raimapu.xlsx", use_container_width=True)
